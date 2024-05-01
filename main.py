import customtkinter

import os
from tkinter import filedialog
from openpyxl import load_workbook

import constant

from PIL import Image, ImageTk
import cv2
from imutils.perspective import four_point_transform
from imutils import contours
import numpy as np
import imutils

class OriginImageFrame(customtkinter.CTkFrame):
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)
        # add widgets onto the frame

class AppButton(customtkinter.CTkButton):
    def __init__(self, master, text, icon, command, width, height):
        super().__init__(master, text=text, text_color=constant.COLOR_BUTTON_TEXT, font=
                                            customtkinter.CTkFont(weight='bold'),
                                            fg_color=constant.COLOR_BUTTON,
                                            hover_color=constant.COLOR_BUTTON_HOVER,
                                            image=icon,
                                            compound='right', width=width, height=height, command=command)

class ScrollableFrameExamFile(customtkinter.CTkScrollableFrame):
    def __init__(self, master, item_list):
        super().__init__(master, width=70, height=5)
        self.grid_columnconfigure(0, weight=1)
        self.item_list=item_list
        self.label_list = []
        for i, item in enumerate(self.item_list):
            self.add_item(item)
    def add_item(self, item):
        label = customtkinter.CTkLabel(self, text=item, padx=5, anchor="w")
        label.grid(row=len(self.label_list), column=0, pady=(0, 10), sticky="w")
        self.label_list.append(label)

class ScrollableFrameCorrectAnswer(customtkinter.CTkScrollableFrame):
    def __init__(self, master, item_list):
        super().__init__(master, width=70, height=40)
        self.grid_columnconfigure(0, weight=1)
        self.label_list = []
        self.item_list = item_list
        for i in range(60):
            self.add_item(i)
    def add_item(self, question_index):
        label = customtkinter.CTkLabel(self, text=f'{question_index+1} - {self.item_list[question_index]}', padx=5, anchor="w")
        label.grid(row=len(self.label_list), column=0, pady=(0, 2), sticky="w")
        self.label_list.append(label)

class Answer:
    def __init__(self, path):
        self.path = path
        self.name_files = []
        # list to store answer alphabet
        self.ANSWER_KEY_LETTER = {}
    def get_answer_indices(self,excel_path):
        # Load the workbook and select the active worksheet
        workbook = load_workbook(excel_path)
        sheet = workbook.active

        # Create a dictionary to map letters to indices
        answer_map = {'A': 0, 'B': 1, 'C': 2, 'D': 3, 'E': 4}

        # Create a list to store the answer indices
        answer_indices = []

        # store answer format letter
        answers_letter = []

        # Assuming the first column has question numbers and the second has answers
        for row in sheet.iter_rows(min_row=1):  # Skip the header row
            answer_letter = row[1].value
            # Convert the letter to an index using the answer_map
            answer_index = answer_map.get(answer_letter, -1)  # Default to -1 if not found
            answer_indices.append(answer_index)
            answers_letter.append(answer_letter)
        return answer_indices, answers_letter

    def read_answers_in_directory(self):
        answers_dict = {}
        ques_index = [i for i in range(60)]
        # Get a list of all files in the directory
        for filename in os.listdir(self.path):
            if filename.endswith(".xlsx"):
                file_path = os.path.join(self.path, filename)
                # Extract the exam code from the filename
                self.name_files.append(filename)
                ma_de = filename.split(".xlsx")[0]
                answer_indices, answers_letter = self.get_answer_indices(file_path)
                answers_dict[ma_de] = answer_indices
                self.ANSWER_KEY_LETTER = dict(zip(ques_index, answers_letter))
        return answers_dict
class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()
        self.geometry('1280x700')
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure(2, weight=1)

        self.init_icon()

        self.origin_image_frame = OriginImageFrame(master=self, width=500, height=550,
                                                         border_color=constant.COLOR_FRAME_BORDER, border_width=2)
        self.origin_image_frame.grid(row=0, column=0, padx=20, pady=20, sticky="nw")

        # init image
        empty_img = customtkinter.CTkImage(light_image=Image.open('assets/icon/select.png'),
                                  dark_image=Image.open('assets/icon/select.png'),
                                  size=(500, 550))
        self.ori_image_label = None

        # add button for choosing folder, image

        # choosing folder button
        self.button_frame = customtkinter.CTkFrame(master=self, width=500, fg_color='transparent')
        self.button_frame.grid(row=1, column=0, padx=20, pady=0, sticky='sw')

        self.choose_folder_button = AppButton(master=self.button_frame, text='Chon folder chua cac ma de',
                                                           icon=self.folder_icon, command=self.choose_folder_command,
                                                            width=240, height=30)
        self.choose_folder_button.grid(row=0, column=0, padx=0, pady=0, sticky='nw')
        self.file_scrollable_frame = ScrollableFrameExamFile(master=self.button_frame, item_list=[])
        self.file_scrollable_frame._scrollbar.configure(height=20)
        self.file_scrollable_frame.grid(row=2, column=0, padx=0, pady=5, sticky='w')

        # choosing image button
        self.choose_image_button = AppButton(master=self.button_frame, text='Chon anh phieu trac nghiem',
                                                           icon=self.photo_icon, command=self.choose_image_command,
                                                            width=240, height=30)
        self.choose_image_button.grid(row=0, column=1, padx=20, pady=0, sticky='ne')

        # add button grade

        self.grade_button = AppButton(master=self, text='Chấm điểm', icon=self.grade_icon, command=self.grade_command,
                                      width=100, height=30)
        self.grade_button.grid(row=0, column=1, padx=0)

        # add graded frame
        self.graded_image_frame = OriginImageFrame(master=self, width=500, height=550,
                                                         border_color=constant.COLOR_FRAME_BORDER, border_width=2)
        self.graded_image_frame.grid(row=0, column=2, padx=20, pady=20, sticky="ne")
        self.graded_image_label = None

        # test_id label
        self.test_id_label = customtkinter.CTkLabel(self, text="Mã đề thi:___", fg_color="transparent",
                                                  font=customtkinter.CTkFont(weight='bold', size=20),
                                                  text_color=constant.COLOR_LABEL_SCORE)
        self.test_id_label.grid(row=1, column=2, padx=20, pady=0, sticky='nw')

        # score label
        self.score_label = customtkinter.CTkLabel(self, text="Điểm:___", fg_color="transparent",
                                                  font=customtkinter.CTkFont(weight='bold', size=20),
                                                  text_color=constant.COLOR_LABEL_SCORE)
        self.score_label.grid(row=2, column=2, padx=20, pady=0, sticky='nw')

        # scrollable correct answers
        self.correct_answers_frame = None

    def choose_folder_command(self):
        directory_path = filedialog.askdirectory()
        answer = Answer(directory_path)
        self.answers_dict = answer.read_answers_in_directory()
        self.answers_dict_letter = answer.ANSWER_KEY_LETTER
        name_files = answer.name_files

        if self.file_scrollable_frame:
            self.file_scrollable_frame.destroy()
        self.file_scrollable_frame = ScrollableFrameExamFile(master=self.button_frame, item_list=name_files)
        self.file_scrollable_frame._scrollbar.configure(height=20)
        self.file_scrollable_frame.grid(row=2, column=0, padx=0, pady=5, sticky='w')

    def choose_image_command(self):
        file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.gif")])
        if file_path:
            self.img_path = file_path
            self.origin_img = cv2.imread(file_path)
            self.origin_img = imutils.resize(self.origin_img, height=1500)
            print(f"Hình ảnh đã chọn: {file_path}")
            image = customtkinter.CTkImage(light_image=Image.open(file_path),
                                  dark_image=Image.open(file_path),
                                  size=(500, 550))
            self.display_ori_img(image, master=self.origin_image_frame)
            if self.graded_image_label:
                self.graded_image_label.destroy()
    def display_ori_img(self, image, master):
        if(self.ori_image_label):
            self.ori_image_label.destroy()
        self.ori_image_label = customtkinter.CTkLabel(master=master, image=image, text="")
        self.ori_image_label.pack()

    def display_graded_img(self, image, master):
        if(self.graded_image_label):
            self.graded_image_label.destroy()
        self.graded_image_label = customtkinter.CTkLabel(master=master, image=image, text="")
        self.graded_image_label.pack()

    def grade_command(self):
        # gray it
        gray = cv2.cvtColor(self.origin_img, cv2.COLOR_BGR2GRAY)
        # blur it
        gray_blur = cv2.GaussianBlur(gray, (5, 5), 0)
        gray_blur_little = cv2.GaussianBlur(gray, (3, 3), 0)
        edged = cv2.Canny(gray_blur, 20, 50)

        # =========================================================================================================

        contours_outer_edge = cv2.findContours(edged.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        contours_outer_edge = imutils.grab_contours(contours_outer_edge)
        sheetCnts = []
        count = 0

        if len(contours_outer_edge) > 2:  # Thay đổi này để đảm bảo rằng có ít nhất 2 đường viền
            contours_outer_edge = sorted(contours_outer_edge, key=cv2.contourArea, reverse=True)
            for c in contours_outer_edge:
                peri = cv2.arcLength(c, True)
                approx = cv2.approxPolyDP(c, 0.05 * peri, True)
                if len(approx) == 4:
                    count += 1
                    sheetCnts.append(approx)
                    if count == 3:  # Thoát vòng lặp sau khi tìm thấy đường viền lớn thứ hai
                        break

        # =========================================================================================================

        exam_code = four_point_transform(self.origin_img, sheetCnts[2].reshape(4, 2))
        warped_code = four_point_transform(gray_blur_little, sheetCnts[2].reshape(4, 2))

        # =========================================================================================================

        # Define the Laplacian kernel for sharpening
        kernel = np.array([
            [0, 1, 0],
            [1, -4, 1],
            [0, 1, 0]
        ])

        # Apply the Laplacian filter for sharpening
        sharpend = cv2.filter2D(warped_code, -1, kernel)
        warped_code = warped_code + sharpend

        thresh_exam_code = cv2.adaptiveThreshold(
            warped_code, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 11, 2)

        # =========================================================================================================

        # Tô các ô đề thi
        cnts = cv2.findContours(thresh_exam_code.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        cnts = imutils.grab_contours(cnts)
        exam_codes = []
        for c in cnts:
            (x, y, w, h) = cv2.boundingRect(c)
            ar = w / float(h)
            if (y >= 20) and (w >= 15 and h >= 15) and (w <= 50 and h <= 50) and ar >= 0.7 and ar <= 1.3:
                exam_codes.append(c)

        # =========================================================================================================

        # Get exam code via painted bubble
        examcodeCnts = contours.sort_contours(exam_codes, method="top-to-bottom")[0]

        examcodeCnts_ordered = []
        for i in np.arange(0, len(examcodeCnts), 3):
            ordered_exam = contours.sort_contours(examcodeCnts[i:i + 3])[0]
            for q in ordered_exam:
                examcodeCnts_ordered.append(q)

        # =========================================================================================================

        code = [0, 0, 0]  # Sử dụng danh sách thay vì chuỗi để lưu trữ các giá trị mã đề
        ordered_code = []
        # Lặp qua các contours của từng ô mã đề
        for (q, i) in enumerate(np.arange(0, len(examcodeCnts_ordered))):
            mask = np.zeros(thresh_exam_code.shape, dtype="uint8")
            cv2.drawContours(mask, examcodeCnts_ordered[i], -1, 255, -1)

            kernel = np.ones((10, 10), np.uint8)
            mask = cv2.dilate(mask, kernel, iterations=1)

            # Áp dụng mask vào ảnh threshold, sau đó đếm số lượng pixel khác không trong vùng bubble
            mask = cv2.bitwise_and(thresh_exam_code, thresh_exam_code, mask=mask)

            total = cv2.countNonZero(mask)
            ordered_code.append(total)

        sorted_indexes = sorted(range(len(ordered_code)), key=lambda i: ordered_code[i], reverse=True)[:3]
        for q in sorted_indexes:
            # Cập nhật giá trị mã đề tương ứng
            if q in [0, 3, 6, 9, 12, 15, 18, 21, 24, 27]:
                code[0] += (int(q / 3) + 1) % 10
            elif q in [1, 4, 7, 10, 13, 16, 19, 22, 25, 28]:
                code[1] += (int(q / 3) + 1) % 10
            elif q in [2, 5, 8, 11, 14, 17, 20, 23, 26, 29]:
                code[2] += (int(q / 3) + 1) % 10
        # Chuyển danh sách code thành chuỗi và in ra
        self.code = int(''.join(str(x) for x in code))
        print(self.code)

        # =========================================================================================================

        paper = four_point_transform(self.origin_img, sheetCnts[0].reshape(4, 2))
        warped = four_point_transform(gray_blur, sheetCnts[0].reshape(4, 2))

        # =========================================================================================================
        thresh = cv2.adaptiveThreshold(
            warped, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 11, 2)
        # vẽ viền cho tất cả để xác nhận lại viền có đúng không
        cnts = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        cnts = imutils.grab_contours(cnts)

        questions = []
        for c in cnts:
            (x, y, w, h) = cv2.boundingRect(c)
            ar = w / float(h)
            if (w >= 15 and h >= 15) and (w <= 50 and h <= 50) and ar >= 0.7 and ar <= 1.3:
                questions.append(c)

        # =========================================================================================================
        answers_dict = self.answers_dict
        ques_index = [i for i in range(60)]
        opts = []
        for ma_de, answer_indices in answers_dict.items():
            if self.code == int(ma_de):
                opts = answer_indices
        ANSWER_KEY = dict(zip(ques_index, opts))

        # =========================================================================================================
        # Get question contours and order
        questionCnts = contours.sort_contours(questions, method="top-to-bottom")[0]

        questionCnts_ordered = []
        for i in np.arange(0, len(questionCnts), 16):
            ordered = contours.sort_contours(questionCnts[i:i + 16])[0]
            for q in ordered:
                questionCnts_ordered.append(q)

        correct = 0
        bubble_thresh = 185
        # Lặp qua các contours của từng câu hỏi
        for (q, i) in enumerate(np.arange(0, len(questionCnts_ordered), 4)):
            # Sắp xếp các contours cho câu hỏi hiện tại từ trái sang phải, sau đó khởi tạo chỉ số của câu trả lời đã được chọn
            cnts = contours.sort_contours(questionCnts_ordered[i:i + 4])[0]
            bubbled = [-1, -1]

            # Lặp qua các contours đã được sắp xếp
            for (j, c) in enumerate(cnts):
                # Tạo mask để chỉ hiển thị "bubbles" cho câu hỏi hiện tại
                mask = np.zeros(thresh.shape, dtype="uint8")
                cv2.drawContours(mask, [c], -1, 255, -1)
                # Áp dụng mask vào ảnh threshold, sau đó đếm số lượng pixel khác không trong vùng bubble
                mask = cv2.bitwise_and(thresh, thresh, mask=mask)
                total = cv2.countNonZero(mask)

                if total > bubbled[0]:
                    bubbled[0] = total
                    bubbled[1] = j

            # Khởi tạo màu của contour và chỉ số của đáp án
            color = (0, 0, 255)
            # Tính xem 4 contours này là answer của câu hỏi số mấy
            row = q // 4;
            col = q % 4
            index_question = 15 * col + row
            k = ANSWER_KEY[index_question]

            # Kiểm tra xem câu trả lời đã được chọn có đúng không
            if k == bubbled[1] and bubbled[0] > bubble_thresh:
                color = (0, 255, 0)
                correct += 1
            # Vẽ viền của câu trả lời đúng trên bài kiểm tra
            cv2.drawContours(paper, [cnts[k]], -1, color, 3)

        # =========================================================================================================
        # số câu tối đa
        max_marks = 60
        # số điểm cho mỗi câu đúng
        positive_marking = 1
        # số điểm cho mỗi câu sai
        # negative_marking = 0

        # grab the test taker
        # Tính toán điểm số
        score = (correct * positive_marking) / max_marks * 100

        # Thông tin về số câu trả lời đúng và điểm số được lưu vào một dictionary info
        info = {
            "correct": correct,
            # "wrong": wrong,
            "score": score
        }

        # # Hiển thị số lượng câu trả lời đúng, sai và điểm số lên trang giấy
        # cv2.putText(paper, "Correct: {}".format(correct), (10, 30),
        #             cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)
        # # cv2.putText(paper, "Wrong: {}".format(wrong), (10, 60),
        # #            cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)
        # cv2.putText(paper, "Score: {:.1f}%".format(score), (180, 30),
        #             cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 0, 0), 2)

        # =========================================================================================================
        img_obj = Image.fromarray(paper)
        result = customtkinter.CTkImage(light_image=img_obj,
                               dark_image=img_obj,
                               size=(500, 550))
        self.display_graded_img(result, master=self.graded_image_frame)
        # =========================================================================================================

        self.test_id_label.configure(text=f'Mã đề thi: {self.code}')
        self.score_label.configure(text=f'Điểm: {score}')
        self.show_correct_answer()

    def show_correct_answer(self):
        self.correct_answers_frame = ScrollableFrameCorrectAnswer(master=self, item_list=self.answers_dict_letter)
        self.correct_answers_frame._scrollbar.configure(height=70)
        self.correct_answers_frame.grid(row=1, column=2, sticky='ne', padx=20, pady=0)

    def init_icon(self):
        self.folder_icon = customtkinter.CTkImage(light_image=Image.open('assets/icon/icons8-folder-50.png'),
                                                 dark_image=Image.open('assets/icon/icons8-folder-50.png'), size=(20, 20))

        self.photo_icon = customtkinter.CTkImage(light_image=Image.open('assets/icon/icon_photo.png'),
                                                  dark_image=Image.open('assets/icon/icon_photo.png'),
                                                  size=(20, 20))
        self.grade_icon = customtkinter.CTkImage(light_image=Image.open('assets/icon/icons8-grade-50.png'),
                                                  dark_image=Image.open('assets/icon/icons8-grade-50.png'),
                                                  size=(20, 20))


if __name__ == "__main__":
    app = App()
    app.configure(fg_color='white')
    app._set_appearance_mode('dark')
    app.title('Chấm điểm trắc nghiệm')
    app.iconbitmap('assets/icon/multiple-choice.ico')
    app.mainloop()